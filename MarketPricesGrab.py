#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Daily Financial Data Scraper

This script allows the user to get financial data from Yahoo finance's historical data page. The data is then
compiled into an Excel spreadsheet where it can be used to compare specific trends and to calculate necessary data.

The Beautiful Soup library is used to open the desired webpage, scrape the topmost line of data from the site,
and store it in a dictionary that has keys matching the column titles from the spreadsheet. The html scraper grabs
from the topmost line using six separate css selectors. The topmost line is the location of the latest data entry
which is updated daily and finalized around 2pm PT each day. The date from the page is also converted into a readable
date format to be inserted into the spreadsheet. The OpenPyXL library is then used to access a spreadsheet in the
computer's memory and update the spreadsheet with the data retrieved from the webpage. The updated spreadsheet is
saved and overwrites the previous version which shouldn't create any issues because the program looks for an empty
row to add the new data to before any edits are made.

The data being inserted is formatted to match the client's format standards (fonts, text size, number formats, borders,
text color) and two equations are included to display the desired information. One of these equations work alongside
manual inputs into the spreadsheet that can be done at a future time.

The program is optimized to run automatically once a day. Therefore it skips days in which the market is closed
when the webpage does not have a new update (weekends and holidays).

If the .xlsx file is open on the device when the script is scheduled to run, an error message will be sent to
remind the user to close the file. Also, if the file location is changed, the file_dir variable must be changed
to match the new file path.
"""

# --- Details --------------------------------------------------------------------------------------------------
__author__ = "Jeremy Bharwani"
__date__ = "12/17/2020"
__license__ = "MIT"
__maintainer__ = "Jeremy Bharwani"
__email__ = "jcb926@gmail.com"
__status__ = "Development"

# --- Built-ins ------------------------------------------------------------------------------------------------
import os
import datetime
import requests

# --- Other Libraries ------------------------------------------------------------------------------------------
import bs4
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side


def main():
    yahoo_url = 'https://finance.yahoo.com/quote/CPSS/history?p=CPSS'
    file_name = 'PriceGrabTest.xlsx'
    file_dir = 'C:\\Users\\Jeremy\\Documents'  # file path not pushed to git
    titles = ['date', 'open', 'high', 'low', 'close', 'volume']
    data = {}

    scrape_data(data, yahoo_url, titles)
    convert_date(data)
    print_to_spreadsheet(data, titles, file_name, file_dir)


# --- High Level Methods ---------------------------------------------------------------------------------------


def scrape_data(data, url, titles):
    """Accesses webpage and fills data dictionary with necessary data

    :param data: dict (str:str)
        Empty dictionary to be filled with keys from titles list and matched with data from webpage
    :param url: str
        Website url, can be changed in main if necessary
    :param titles: list (str)
        List of each column's title, each in string format
    """
    # list of each css selector position, retrieved from manually inspecting website's html script
    soup_location = [r'#Col1-1-HistoricalDataTable-Proxy > section > div.Pb\(10px\).Ovx\(a\).W\(100\%\) > table > '
                     r'tbody > tr:nth-child(1) > td.Py\(10px\).Ta\(start\).Pend\(10px\) > span',
                     r'#Col1-1-HistoricalDataTable-Proxy > section > div.Pb\(10px\).Ovx\(a\).W\(100\%\) > table > '
                     r'tbody > tr:nth-child(1) > td:nth-child(2) > span',
                     r'#Col1-1-HistoricalDataTable-Proxy > section > div.Pb\(10px\).Ovx\(a\).W\(100\%\) > table > '
                     r'tbody > tr:nth-child(1) > td:nth-child(3) > span',
                     r'#Col1-1-HistoricalDataTable-Proxy > section > div.Pb\(10px\).Ovx\(a\).W\(100\%\) > table > '
                     r'tbody > tr:nth-child(1) > td:nth-child(4) > span',
                     r'#Col1-1-HistoricalDataTable-Proxy > section > div.Pb\(10px\).Ovx\(a\).W\(100\%\) > table > '
                     r'tbody > tr:nth-child(1) > td:nth-child(5) > span',
                     r'#Col1-1-HistoricalDataTable-Proxy > section > div.Pb\(10px\).Ovx\(a\).W\(100\%\) > table > '
                     r'tbody > tr:nth-child(1) > td:nth-child(7) > span']

    # accesses specified website's text
    webpage = requests.get(url)
    soup = bs4.BeautifulSoup(webpage.text, 'html.parser')

    # fill data dictionary with data from website
    for category in range(len(titles)):
        grab = soup.select(soup_location[category])
        data[titles[category]] = grab[0].text.strip()


def convert_date(data):
    """Converts the date format from the webpage into an excel friendly format

    :param data: dict (str:str)
        Contains values scraped from website with keys matching the column titles
    """
    date = data['date']
    year = int(date[8:12])
    day = int(date[4:6])
    month_num = datetime.datetime.strptime(date[0: 3], "%b")  # converts 3 letter month code to int

    converted_date = datetime.datetime(year, month_num.month, day)
    data['date'] = converted_date


def print_to_spreadsheet(data, titles, file_name, file_dir):
    """Opens spreadsheets, updates data, and saves the updated version

    :param data: dict (str:str)
        Contains values scraped from website with keys matching the column titles which are found in
        the titles list
    :param titles: list (str)
        List of each column's title, each in string format
    :param file_name: str
        Name of the Excel spreadsheet that is being updated, can be changed in main if necessary
    :param file_dir: str
        Directory in which the Excel spreadsheet is saved (overwrites previous save but should not
        delete anything, only adds to previous versions), can be changed in main if necessary
    """
    columns = ['A', 'B', 'C', 'D', 'E', 'F']
    confirm = True
    os.chdir(file_dir)

    # opens workbook if it is in the correct path and navigates to the next empty row
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        print('Could not open file\nConfirm file path has not changed from:')
        print(os.getcwd())
        raise  # ends program with a possible error explanation
    sheet = workbook['volume limit']
    row = next_empty_row(sheet)
    draw_borders(workbook, file_name, sheet, row, columns)

    # inserts converted values into the spreadsheet
    for num in range(len(titles)):
        cell = columns[num] + str(row)
        if num == 0:
            check_market_open(data, sheet, row)
            sheet[cell].value = data['date']
            sheet[cell].number_format = u'mm/dd/yyyy'  # date format
        elif num == (len(titles) - 1):
            vol = data['volume']
            vol_num = vol.replace(',', '')
            sheet[cell].value = int(vol_num)
            sheet[cell].number_format = u'#,##0'  # volume format
        else:
            sheet[cell].value = float(data[titles[num]])
            sheet[cell].number_format = u'0.00'  # format for all other values
        sheet[cell].font = Font(name='Arial', size=11)
    calculate_condition_limit(sheet, row)
    calculate_violation(sheet, row)

    # saves workbook and avoids error message if file is still open
    try:
        workbook.save(file_name)
    except PermissionError:
        print('Could not update, the file must be closed on this device')
        confirm = False

    if confirm:
        print('File Update Confirmation')


# --- Low Level Methods ---------------------------------------------------------------------------------------


def next_empty_row(sheet):
    """Searches for the next empty row in the spreadsheet starting from the top

    :param sheet: list
        Specified spreadsheet that this method is searching within
    :return: int
        Row number of the first empty row from the top
    """
    row = 3  # skips first two rows which contain titles
    while sheet.cell(row=row, column=1).value is not None:
        row += 1

    return row


def draw_borders(workbook, file_name, sheet, row, columns):
    """Draws a border beneath the current data at the end of the week

    :param workbook: obj
        The spreadsheet that was opened by the program and can be overwritten with new data
    :param file_name: str
        Name of the Excel spreadsheet that is being updated, can be changed in main if necessary
    :param sheet: list
        Specified spreadsheet that this method is editing and searching within
    :param row: int
        Row number of the first empty row from the top
    :param columns: list
        Contains the letter values of all the columns that are used within the sheet
    """
    today = datetime.datetime.now()
    row -= 1
    if today.strftime("%a") == 'Sat':
        bottom_border = Border(bottom=Side(style='thin'))
        for letter in columns:
            sheet.cell(row=row, column=letter).border = bottom_border

        # has its own save because the normal save is never reached on a saturday
        try:
            workbook.save(file_name)
        except PermissionError:
            print('Could not update, the file must be closed on this device')


def check_market_open(data, sheet, row):
    """Exits program if the date from the webpage is the same as the last entry in the spreadsheet

    :param data: dict (str:str)
        Contains values scraped from website with keys matching the column titles which are found in
        the titles list
    :param sheet: dict (str:any)
        Contains the details of every cell in the spreadsheet, can call '.value' to get specific cell's contents
    :param row: int
        Row number of the first empty row from the top
    """
    if sheet.cell(row=row - 1, column=1).value == data['date']:
        print('Markets are closed today. No update.')
        exit()


def calculate_condition_limit(sheet, row):
    """Inserts formula that is dependent on a certain set of previously inserted values

    :param sheet: list
        Specified spreadsheet that this method is editing and searching within
    :param row: int
        Row number of the first empty row from the top
    """
    cell = 'G' + str(row)
    start_row = 0
    end_row = 0
    border_count = 0

    # searches for cells for the previous four weeks which vary in number of days depending on holidays
    while border_count < 5:
        row -= 1
        if 'thin' in str(sheet.cell(row=row, column=1).border):
            border_count += 1
            if border_count == 1:
                end_row = row
            elif border_count == 5:
                start_row = row + 1

    sheet[cell].value = f"=ROUND(AVERAGE($F${start_row}:$F${end_row})*0.25,-2)"
    sheet[cell].number_format = u'#,##0'
    sheet[cell].font = Font(name='Arial', size=11)


def calculate_violation(sheet, row):
    """Inserts formula that can interact with future user input into the spreadsheet

    :param sheet: list
        Specified spreadsheet that this method is editing and searching within
    :param row: int
        Row number of the first empty row from the top
    """
    cell = 'I' + str(row)
    sheet[cell].value = f'=IF(H{row}<G{row},"",+H{row}-G{row})'
    sheet[cell].number_format = u'[Red]#,##0'
    sheet[cell].font = Font(name='Arial', size=11)


if __name__ == '__main__':
    main()
